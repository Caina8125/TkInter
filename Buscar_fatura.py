import pandas as pd
import pyautogui
import time
from abc import ABC
from tkinter import filedialog
import tkinter.messagebox
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from seleniumwire import webdriver


class PageElement(ABC):
    def __init__(self, webdriver, url=''):
        self.webdriver = webdriver
        self.url = url
    def open(self):
        self.webdriver.get(self.url)

class Login(PageElement):
    multiusuario = (By.XPATH, '/html/body/div[3]/div[3]/div/form/div[1]/label')
    prestador = (By.XPATH, '//*[@id="login_code"]')
    cpf = (By.XPATH, '//*[@id="login_cpf"]')
    senha = (By.XPATH, '//*[@id="login_password"]')
    logar = (By.XPATH, '//*[@id="btnLogin"]')

    def exe_login(self, prestador, cpf, senha):
        self.webdriver.find_element(*self.multiusuario).click()
        self.webdriver.find_element(*self.prestador).send_keys(prestador)
        self.webdriver.find_element(*self.cpf).send_keys(cpf)
        self.webdriver.find_element(*self.senha).send_keys(senha)
        self.webdriver.find_element(*self.logar).click()
        time.sleep(4)

class caminho(PageElement):
    Alerta = (By.XPATH, '/html/body/div[2]/div/center/a')

    def exe_caminho(self):
        try:
            self.webdriver.find_element(*self.Alerta).click()
        except:
            print('Não tem alerta')

        driver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")
        time.sleep(3)

class capturar_protocolo(PageElement):
    inserir_protocolo = (By.XPATH, '//*[@id="NroProtocolo"]')
    baixar = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[2]/article/form/div/a')
    elemento2 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[5]')
    elemento3 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[3]/td[5]')

    def exe_capturar(self):
        global count
        count = 0
        
        faturas_df = pd.read_excel(planilha)
        for index, linha in faturas_df.iterrows():
            
            count = count + 1

            protocolo_plan =  f"{linha['Protocolo']}".replace(".0","")
            fatura_plan =  f"{linha['Faturas']}".replace(".0","")
            if ((f"{linha['Verificação']}" == "Fatura encontrada") or (protocolo_plan == "Total Geral")):
                print(count,')',protocolo_plan, ": Fatura encontrada =>", fatura_plan)
                continue
            
            print(count,')','Buscanco a fatura do Protocolo =>', protocolo_plan )
            driver.find_element(*self.inserir_protocolo).send_keys(protocolo_plan)
            driver.find_element(*self.baixar).click()
            time.sleep(0.7)

            #Bloco de código que insere o número da fatura na planilha
            fatura_site = driver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[3]').text
            n_fatura = [fatura_site]
            df = pd.DataFrame(n_fatura)
            book = load_workbook(planilha)
            writer = pd.ExcelWriter(planilha, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'Faturas', startrow = count, startcol = 1, header=False, index=False)
            writer.save()

            capturar_protocolo(webdriver,url).confere()

            driver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")

    def confere(self):

        plan = planilha
        plan_atualizada = pd.read_excel(plan)
        dados = pd.DataFrame(plan_atualizada)
        dados = dados.iloc[count - 1]
        try:
            fatura_plan = dados["Faturas"].astype(str).replace(".0","")
        except:
            fatura_plan = str(dados["Faturas"])
        try:
            protocolo_plan = dados["Protocolo"].astype(str).replace(".0","")
        except:
            protocolo_plan = str(dados["Protocolo"])

        fatura_site = driver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[3]').text
        protocolo_site  = driver.find_element(By.XPATH,'//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[1]').text

        if ((protocolo_plan == protocolo_site) & (fatura_site == fatura_plan)):
            confere = ["Fatura encontrada"]
            df = pd.DataFrame(confere)
            book = load_workbook(plan)
            writer = pd.ExcelWriter(plan, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'Faturas', startrow=count, startcol=5, header=False, index=False)
            writer.save()
        else:
            erro = ["Verificar"]
            df = pd.DataFrame(erro)
            book = load_workbook(plan)
            writer = pd.ExcelWriter(plan, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'Faturas', startrow= count, startcol=5, header=False, index=False)
            writer.save()

#-------------------------------------------------------------------------------------------------------------------------

def iniciar():
    url = 'https://www2.geap.com.br/auth/prestador.asp'
    # planilha = filedialog.askopenfilename()

    options = {
        'proxy' : {
            'http': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128',
            'https': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128'
        }
    }

    driver = webdriver.Chrome(seleniumwire_options= options)

    driver.maximize_window()

    login_page = Login(driver , url)
    
    login_page.open()
   
    login_page.exe_login(
        prestador = "23003723",
        cpf = '66661692120',
        senha = "amhpdf0073"
    )

    caminho(webdriver, url).exe_caminho()
    # try:
    capturar_protocolo(webdriver, url).exe_capturar()

    tkinter.messagebox.showinfo( 'Automação GEAP Financeiro' , 'Busca de Faturas na GEAP Concluído 😎✌' )

    # except:
    #     tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal da GEAP caiu 😢' )

#-------------------------------------------------------------------------------------------------------------------------------------------------------------


